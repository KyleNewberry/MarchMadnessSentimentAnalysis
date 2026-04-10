import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

API_KEY = ""
MAX_COMMENTS = None  # Change this to fetch more (or set to None for all)
ORDER = "relevance"  # "relevance" for top comments, "time" for newest first
video_list = ["a0P4WeUURks", "jZXyRxUWuG4", "f1RE8muF2J8", "CIa6QieqGJ4", "MVrolABrKuI", "k4eHiWmGYQk", "UuAYHMg6MYU", "-FUy9tblgSQ", "fEGwbBsKTNU", "8sC-92iyG5Y", "Y56y1vpZDMU", "_lpUpaDvMfY", "jQ1oxzALkdw", "B3rSJMyoE08", "1RY_UbM3BeY", "BCOixn9ohb0", "516hquMoDyI", "eqQ7wSYhhGg", "B9dLynEr5cg", "5zcKvmpqToo", "qjkE3suanAY", "DNOWFBMNrWU", "Okgp2Nq2Hko", "-i2fw_fc_k4", "wggZpEfnPlY", "5uNCGloqTbk", "TS-_20We18I", "Lw77zuTPapc", "pIHDY0TtLQA", "ubnicqeZvAo", "qwx0yWsfURA", "zytc0_Ye7Yo", "DE6uE_Lu3Ns", "_1J0iu9H4VE", "2o_hGvmMNkM", "-_Cxj-W3p5k", "2ItzwrfLnp0", "ydPqbnRlI8E", "mD5KUn5Q4Sk", "qP1J6STlrl8", "lQQ7x05Ufr8", "vZBvf_41iFU", "zcsDIi5nuZA", "xLLDeGErwTs", "lAPLJ1tugZs", "KOiSk1hr16Y", "ye8-D2zXns4", "A78xdvc9VmI", "rVgdMl2qs8I", "a3XeaZ2TFig", "aFr9ErWOO54", "FwLknx-BXaA", "JvxsFtlnCUc", "GH4FegFqNTs", "s8MTiqZ7rFc", "xWPjxVlb8mk", "tNyN_y1bVaI", "uF_59KHDmlM", "M6nHvwW5VlA", "R48Mxg_UhIA", "rB9KGU2UI2A", "-OmxGRrPry4", "2ZLXQ62lWlQ"]
name_list = ["UCONN-Duke", "Michigan-Tennessee", "Arizona-Purdue", "Illinois-Iowa", "UCONN-Michigan St.", "Duke-St. Johns", "Michigan-Alabama", "Arizona-Arkansas", "Tennessee-Iowa St.", "Illinois-Houston", "Purdue-Texas", "Iowa-Nebraska", "Texas-Gonzaga", "UCONN-UCLA", "Iowa-Florida", "St. Johns-Kansas", "Iowa St.-Kentucky", "Tennessee-Virginia", "Arkansas-High Point", "Nebraska-Vanderbilt", "Michigan St.-Louisville", "Michigan-St. Louis", "Duke-TCU", "Alabama-Texas Tech", "Arizona-Utah St.", "Houston-TexasA&M", "Illinois-VCU", "Purdue-Miami(FL)", "Duke-Siena", "UCLA-UCF", "TCU-Ohio St.", "St. Johns-UNI", "Kansas-CBU", "Lousiville-USF", "Michigan St.-North Dakota State", "UCONN-Furman", "Florida-Prairie View A&M", "Iowa-Clemson", "Vanderbilt-McNeese", "Nebraska-Troy", "VCU-UNC", "Illinois-Penn", "TexasA&M-St. Mary's", "Houston-Idaho", "Michigan-Howard", "St. Louis-Georgia", "Texas Tech-Akron", "Alabama-Hofstra", "Tennessee-Miami(OH)", "Virginia-Wright St.", "Kentucky-Santa Clara", "Iowa St.-Tennessee St.", "Arizona-LIU", "Utah St.-Villanova", "High Point-Wisconsin", "Arkansas-Hawaii", "Texas-BYU", "Gonzaga-Kennesaw St.", "Miami-Missouri", "Purdue-Queens", "UCONN-Illinois", "Michigan-Arizona", "Michigan-UCONN"]

def fetch_comments(api_key, video_id, max_comments, order):
    comments = []
    page_token = None
    url = "https://www.googleapis.com/youtube/v3/commentThreads"

    while True:
        params = {
            "part": "snippet",
            "videoId": video_id,
            "maxResults": 100,
            "order": order,
            "key": api_key,
        }
        if page_token:
            params["pageToken"] = page_token

        response = requests.get(url, params=params)
        data = response.json()

        if "error" in data:
            raise Exception(f"API Error: {data['error']['message']}")

        for item in data.get("items", []):
            s = item["snippet"]["topLevelComment"]["snippet"]
            comments.append({
                "author": s["authorDisplayName"],
                "comment": s["textDisplay"],
                "likes": s["likeCount"],
                "date": s["publishedAt"][:10],
            })
            print(f"  [{len(comments)}] {s['authorDisplayName']}: {s['textDisplay'][:60]}...")

        page_token = data.get("nextPageToken")
        if not page_token:
            break
        if max_comments and len(comments) >= max_comments:
            break

    return comments[:max_comments] if max_comments else comments


def save_to_excel(comments, video_id, output_file):
    wb = openpyxl.Workbook()

    # --- Comments sheet ---
    ws = wb.active
    ws.title = "Comments"

    headers = ["#", "Author", "Comment", "Likes", "Date"]
    header_fill = PatternFill("solid", start_color="D9E1F2")
    header_font = Font(bold=True, name="Arial")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, c in enumerate(comments, 1):
        ws.append([i, c["author"], c["comment"], c["likes"], c["date"]])

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 120
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 14

    ws.freeze_panes = "A2"

    # --- Info sheet ---
    info_ws = wb.create_sheet("Info")
    info_data = [
        ["Video ID", video_id],
        ["Video URL", f"https://www.youtube.com/watch?v={video_id}"],
        ["Total Comments", len(comments)],
        ["Date Exported", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
    ]
    for row in info_data:
        info_ws.append(row)
        info_ws.cell(row=info_ws.max_row, column=1).font = Font(bold=True, name="Arial")

    info_ws.column_dimensions["A"].width = 18
    info_ws.column_dimensions["B"].width = 50

    wb.save(output_file)
    print(f"\nSaved {len(comments)} comments to '{output_file}'")


if __name__ == "__main__":
    for ID in range(len(video_list)):
        VIDEO_ID = video_list[ID]
        print(f"Fetching comments for video: {name_list[ID]}")
        comments = fetch_comments(API_KEY, VIDEO_ID, MAX_COMMENTS, ORDER)
        save_to_excel(comments, VIDEO_ID, f"School/SentimentAnalysis/ExcelSheets/{name_list[ID]}_comments.xlsx")