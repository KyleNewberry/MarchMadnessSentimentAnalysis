"""
March Madness Sentiment Analysis  —  FINAL VERSION
=====================================================
- Reads team data from MarchMadnessAliases.json
- Detects which team each comment is about
- Runs Cardiff RoBERTa sentiment model
- Produces 3 output Excel files:
    1. all_comments_labeled.xlsx  — every comment tagged + scored
    3. sentiment_summary.xlsx     — sentiment breakdown per team per game

SETUP: Place this script + MarchMadnessAliases.json in the same folder
       as all your game .xlsx files, then run:  python sentiment_analysis.py
"""

import json, re, os, glob
import pandas as pd
import numpy as np
from scipy.special import softmax
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from transformers import AutoTokenizer, AutoConfig, AutoModelForSequenceClassification

# ─── CONFIGURE THESE ────────────────────────────────────────────────────────
INPUT_FILES  = "*.xlsx"
COMMENT_COL  = "Comment"
SHEET_NAME   = 0                          # 0 = first sheet
TEAMS_JSON   = "MarchMadnessAliases.json"
# ────────────────────────────────────────────────────────────────────────────

# Output filenames — these are always skipped as inputs so we never re-process them
OUTPUT_FILES = {
    "all_comments_labeled.xlsx",
    "sentiment_summary.xlsx",
}

# ── Styling ──────────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
TEAM1_FILL   = PatternFill("solid", fgColor="D9E1F2")   # light blue
TEAM2_FILL   = PatternFill("solid", fgColor="FCE4D6")   # light orange
BOTH_FILL    = PatternFill("solid", fgColor="E2EFDA")   # light green
NEITHER_FILL = PatternFill("solid", fgColor="F2F2F2")   # light grey
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT    = Font(name="Arial", size=10)
THIN_BORDER  = Border(left=Side(style="thin"), right=Side(style="thin"),
                      top=Side(style="thin"),  bottom=Side(style="thin"))
SENTIMENT_COLORS = {
    "Positive": PatternFill("solid", fgColor="C6EFCE"),
    "Neutral":  PatternFill("solid", fgColor="FFEB9C"),
    "Negative": PatternFill("solid", fgColor="FFC7CE"),
}


# ════════════════════════════════════════════════════════════════════════════
# HELPERS
# ════════════════════════════════════════════════════════════════════════════

def load_teams(json_path):
    with open(json_path, "r", encoding="utf-8") as f:
        raw = json.load(f)
    return {
        name: {
            "aliases":      [a.lower() for a in d.get("aliases", [])],
            "coach":        [c.lower() for c in d.get("coach", [])],
            "players":      [p.lower() for p in d.get("players", [])],
            "ignore_words": [i.lower() for i in d.get("ignore_words", [])],
        }
        for name, d in raw.items()
    }


def detect_teams_from_filename(filename, teams):
    """Parse 'Alabama-Texas_Tech_comments.xlsx' → ['Alabama Crimson Tide', 'Texas Tech']"""
    base = re.sub(r'_?comments$', '',
                  os.path.splitext(os.path.basename(filename))[0],
                  flags=re.IGNORECASE)
    parts   = base.split("-")
    t1r = parts[0].replace("_", " ").strip().lower() if parts      else ""
    t2r = parts[1].replace("_", " ").strip().lower() if len(parts)>1 else ""

    found = []
    for name, data in teams.items():
        for alias in data["aliases"]:
            if alias in t1r or alias in t2r or t1r in alias or t2r in alias:
                if name not in found:
                    found.append(name)
                break

    # Sort so team1 from filename comes first
    if len(found) == 2:
        found.sort(key=lambda t: 0 if any(
            a in t1r or t1r in a for a in teams[t]["aliases"]) else 1)

    return found if found else [t1r.title(), t2r.title()]


def comment_matches_team(text_lower, data):
    """Return True if the comment contains any alias/coach/player keyword."""
    cleaned = text_lower
    for iw in data["ignore_words"]:
        cleaned = re.sub(r'(?<!\w)' + re.escape(iw) + r'(?!\w)', ' ', cleaned)
    for kw in data["aliases"] + data["coach"] + data["players"]:
        if re.search(r'(?<!\w)' + re.escape(kw) + r'(?!\w)', cleaned):
            return True
    return False


def detect_team(text, t1n, t1d, t2n, t2d):
    tl = str(text).lower()
    h1 = comment_matches_team(tl, t1d)
    h2 = comment_matches_team(tl, t2d)
    if h1 and h2: return "Both Teams"
    if h1:        return t1n
    if h2:        return t2n
    return "Neither"


def preprocess(text):
    return " ".join(
        "@user" if t.startswith("@") and len(t) > 1 else
        "http"  if t.startswith("http") else t
        for t in str(text).split()
    )


def get_sentiment(text, tokenizer, model, config):
    encoded = tokenizer(preprocess(text), return_tensors="pt",
                        truncation=True, max_length=512)
    scores  = softmax(model(**encoded)[0][0].detach().numpy())
    # .title() ensures "positive"->"Positive", "POSITIVE"->"Positive", etc.
    label   = config.id2label[int(np.argmax(scores))].title()
    return label, round(float(scores[0]),4), round(float(scores[1]),4), round(float(scores[2]),4)


def style_header(ws, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def auto_widths(ws, max_w=60):
    for col in ws.columns:
        w = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w+4, max_w)


def write_sheet(ws, df, row_fill=None):
    headers = list(df.columns)
    ws.append(headers)
    style_header(ws, len(headers))
    for _, row in df.iterrows():
        ws.append(list(row))
        r = ws.max_row
        for ci, col_name in enumerate(headers, 1):
            cell = ws.cell(row=r, column=ci)
            cell.font = BODY_FONT; cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=(col_name == COMMENT_COL), vertical="top")
            if col_name == "sentiment" and cell.value in SENTIMENT_COLORS:
                cell.fill = SENTIMENT_COLORS[cell.value]
            elif row_fill:
                cell.fill = row_fill
    ws.row_dimensions[1].height = 28
    auto_widths(ws)
    ws.freeze_panes = "A2"


def make_sent_summary(df, group_cols):
    """Build a sentiment pivot table with counts and percentages."""
    # Normalise sentiment to Title Case before grouping
    df = df.copy()
    df["sentiment"] = df["sentiment"].astype(str).str.title()

    grp = df.groupby(group_cols + ["sentiment"]).size().unstack(fill_value=0).reset_index()

    # Print what columns came out of the unstack so we can debug if needed
    print(f"    [summary] columns after unstack: {list(grp.columns)}")

    for c in ["Positive", "Neutral", "Negative"]:
        if c not in grp.columns:
            grp[c] = 0

    keep = group_cols + [c for c in ["Positive","Neutral","Negative"] if c in grp.columns]
    grp  = grp[keep].copy()
    grp["Total"]      = grp[["Positive","Neutral","Negative"]].sum(axis=1)
    denom = grp["Total"].replace(0, np.nan)
    grp["% Positive"] = (grp["Positive"] / denom * 100).round(1).fillna(0)
    grp["% Neutral"]  = (grp["Neutral"]  / denom * 100).round(1).fillna(0)
    grp["% Negative"] = (grp["Negative"] / denom * 100).round(1).fillna(0)
    return grp


# ════════════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════════════
def main():
    # -- Validate JSON --
    if not os.path.exists(TEAMS_JSON):
        print(f"ERROR: '{TEAMS_JSON}' not found in this folder.")
        print("Make sure MarchMadnessAliases.json is in the same folder as this script.")
        return

    teams = load_teams(TEAMS_JSON)
    print(f"Loaded {len(teams)} teams from {TEAMS_JSON}")

    # -- Load sentiment model --
    MODEL = "cardiffnlp/twitter-roberta-base-sentiment-latest"
    print("Loading sentiment model... (first run downloads ~500MB, then it's cached)")
    tokenizer = AutoTokenizer.from_pretrained(MODEL)
    config    = AutoConfig.from_pretrained(MODEL)
    model     = AutoModelForSequenceClassification.from_pretrained(MODEL)
    print("Model ready!\n")

    # -- Find input files (never include our own outputs) --
    input_files = [
        f for f in sorted(glob.glob(INPUT_FILES))
        if os.path.basename(f) not in OUTPUT_FILES
    ]
    if not input_files:
        print("No input .xlsx files found. Place your game files in this folder.")
        return
    print(f"Found {len(input_files)} game file(s).\n")

    # -- Process each game file --
    all_rows = []
    for filepath in input_files:
        print(f"Processing: {filepath}")
        try:
            df = pd.read_excel(filepath, sheet_name=SHEET_NAME)
        except Exception as e:
            print(f"  ERROR reading file: {e} — skipping.\n")
            continue

        if COMMENT_COL not in df.columns:
            print(f"  WARNING: column '{COMMENT_COL}' not found — skipping.\n")
            continue

        gt   = detect_teams_from_filename(filepath, teams)
        t1n  = gt[0] if len(gt) >= 1 else "Team 1"
        t2n  = gt[1] if len(gt) >= 2 else "Team 2"
        t1d  = teams.get(t1n, {"aliases":[],"coach":[],"players":[],"ignore_words":[]})
        t2d  = teams.get(t2n, {"aliases":[],"coach":[],"players":[],"ignore_words":[]})
        game = re.sub(r'_?comments$', '',
                      os.path.splitext(os.path.basename(filepath))[0],
                      flags=re.IGNORECASE).replace("_", " ")

        print(f"  Teams detected : {t1n}  vs  {t2n}")
        print(f"  Comments to process: {len(df)}")

        for _, row in df.iterrows():
            comment  = row.get(COMMENT_COL, "")
            team_tag = detect_team(comment, t1n, t1d, t2n, t2d)
            sentiment, neg, neu, pos = get_sentiment(comment, tokenizer, model, config)
            all_rows.append({
                "game":           game,
                "team_1":         t1n,
                "team_2":         t2n,
                **{k: row[k] for k in df.columns if k != COMMENT_COL},
                COMMENT_COL:      comment,
                "team_tag":       team_tag,
                "sentiment":      sentiment,
                "score_negative": neg,
                "score_neutral":  neu,
                "score_positive": pos,
            })
        print(f"  Done.\n")

    if not all_rows:
        print("No comments were processed. Check your file and column names.")
        return

    master = pd.DataFrame(all_rows)

    # ── DIAGNOSTIC: always print this so you can verify sentiment is working ──
    print("── Team Tag Distribution ───────────────────────────")
    print(master["team_tag"].value_counts().to_string())
    print("\n── Sentiment Distribution ──────────────────────────")
    print(master["sentiment"].value_counts().to_string())
    print(f"\n── Sample sentiment values (first 5): {master['sentiment'].head().tolist()}")
    print(f"── Sentiment dtype: {master['sentiment'].dtype}")
    print()

    # Split into 4 buckets
    df_t1      = master[master["team_tag"] == master["team_1"]].reset_index(drop=True)
    df_t2      = master[master["team_tag"] == master["team_2"]].reset_index(drop=True)
    df_both    = master[master["team_tag"] == "Both Teams"].reset_index(drop=True)
    df_neither = master[master["team_tag"] == "Neither"].reset_index(drop=True)

    # ── OUTPUT 1: all_comments_labeled.xlsx ─────────────────────────────
    print("Writing all_comments_labeled.xlsx ...")
    wb1 = Workbook(); ws1 = wb1.active; ws1.title = "All Comments"
    headers = list(master.columns)
    ws1.append(headers); style_header(ws1, len(headers))

    tag_fill = {"Both Teams": BOTH_FILL, "Neither": NEITHER_FILL}
    for _, row in master.iterrows():
        ws1.append(list(row)); r = ws1.max_row
        tag = row["team_tag"]
        if   tag == row["team_1"]: rf = TEAM1_FILL
        elif tag == row["team_2"]: rf = TEAM2_FILL
        else:                      rf = tag_fill.get(tag, NEITHER_FILL)
        for ci, col_name in enumerate(headers, 1):
            cell = ws1.cell(row=r, column=ci)
            cell.font = BODY_FONT; cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=(col_name==COMMENT_COL), vertical="top")
            if col_name == "sentiment" and cell.value in SENTIMENT_COLORS:
                cell.fill = SENTIMENT_COLORS[cell.value]
            else:
                cell.fill = rf

    ws1.row_dimensions[1].height = 28; auto_widths(ws1); ws1.freeze_panes = "A2"

    wl = wb1.create_sheet("Legend")
    legend = [
        ("Blue rows",    "Comments about Team 1"),
        ("Orange rows",  "Comments about Team 2"),
        ("Green rows",   "Comments mentioning both teams"),
        ("Grey rows",    "Neither team mentioned"),
        ("", ""),
        ("Green cell",   "Positive sentiment"),
        ("Yellow cell",  "Neutral sentiment"),
        ("Red cell",     "Negative sentiment"),
    ]
    lfills = [TEAM1_FILL, TEAM2_FILL, BOTH_FILL, NEITHER_FILL, None,
              SENTIMENT_COLORS["Positive"], SENTIMENT_COLORS["Neutral"], SENTIMENT_COLORS["Negative"]]
    for i, (a, b) in enumerate(legend, 1):
        wl.cell(i,1,a).font = Font(name="Arial", size=10)
        wl.cell(i,2,b).font = Font(name="Arial", size=10)
        if i-1 < len(lfills) and lfills[i-1]:
            wl.cell(i,1).fill = lfills[i-1]; wl.cell(i,2).fill = lfills[i-1]
    wl.column_dimensions["A"].width = 28; wl.column_dimensions["B"].width = 38
    wb1.save("all_comments_labeled.xlsx")
    print("  Saved: all_comments_labeled.xlsx")

    # ── OUTPUT 3: sentiment_summary.xlsx ────────────────────────────────
    print("Writing sentiment_summary.xlsx ...")
    wb3 = Workbook(); wb3.remove(wb3.active)

    # Full breakdown
    ws_full = wb3.create_sheet("By Game & Team")
    full = make_sent_summary(master, ["game", "team_tag"])
    write_sheet(ws_full, full)
    col_names = list(full.columns)
    for r in range(2, ws_full.max_row+1):
        for cn, fill in [("Positive", SENTIMENT_COLORS["Positive"]),
                         ("Neutral",  SENTIMENT_COLORS["Neutral"]),
                         ("Negative", SENTIMENT_COLORS["Negative"])]:
            if cn in col_names:
                ws_full.cell(row=r, column=col_names.index(cn)+1).fill = fill

    # Per-bucket sentiment sheets
    for sheet_name, df_s, gcols, rename_col, fill in [
        ("Team 1 Sentiment", df_t1,      ["game","team_1"], "team_1", TEAM1_FILL),
        ("Team 2 Sentiment", df_t2,      ["game","team_2"], "team_2", TEAM2_FILL),
        ("Both Sentiment",   df_both,    ["game"],          None,     BOTH_FILL),
        ("Neither Sentiment",df_neither, ["game"],          None,     NEITHER_FILL),
    ]:
        ws = wb3.create_sheet(sheet_name)
        if df_s.empty:
            ws.append([f"No comments in this category."])
            continue
        df_s2 = df_s.copy()
        if rename_col: df_s2 = df_s2.rename(columns={rename_col: "team"})
        gcols2 = ["game","team"] if rename_col else ["game"]
        s = make_sent_summary(df_s2, gcols2)
        write_sheet(ws, s, row_fill=fill)

    wb3.save("sentiment_summary.xlsx")
    print("  Saved: sentiment_summary.xlsx")

    print("""
All done! Three output files created:
  • all_comments_labeled.xlsx  — every comment tagged and color-coded
  • sentiment_summary.xlsx     — sentiment counts and percentages
""")


if __name__ == "__main__":
    main()